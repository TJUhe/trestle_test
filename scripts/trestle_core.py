from __future__ import annotations

import csv
import json
import math
from collections import Counter
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Iterable


Point3 = tuple[float, float, float]


def parse_float_list(value: object) -> list[float]:
    if value in (None, ""):
        return []
    if isinstance(value, (list, tuple)):
        return [float(item) for item in value if item not in (None, "")]
    text = str(value).strip()
    for token in ["，", "；", ";", "|", "\n", "\t"]:
        text = text.replace(token, ",")
    text = text.replace("、", ",")
    return [float(item.strip()) for item in text.split(",") if item.strip()]


def parse_float_range(value: object) -> tuple[float, float]:
    if isinstance(value, (list, tuple)) and len(value) >= 2:
        return (float(value[0]), float(value[1]))
    text = str(value).strip()
    for token in ["~", "至", "到", "，", ",", ";", "；", "|"]:
        text = text.replace(token, "-")
    parts = [item.strip() for item in text.split("-") if item.strip()]
    if len(parts) < 2:
        raise ValueError(f"Range value needs two numbers: {value!r}")
    return (float(parts[0]), float(parts[1]))


def spans_from_stations(stations: list[float]) -> list[float]:
    if not stations:
        return []
    ordered = [float(item) for item in stations]
    if abs(ordered[0]) < 1e-9:
        return [0.0, *[b - a for a, b in zip(ordered[:-1], ordered[1:])]]
    return [ordered[0], *[b - a for a, b in zip(ordered[:-1], ordered[1:])]]


LAYER_TABLE: dict[str, dict[str, object]] = {
    "deck_main_beam": {"cad_layer": "ZJ8_DECK_MAIN", "aci": 4, "section": "HM-220x220", "role": "平台主梁"},
    "deck_cross_beam": {"cad_layer": "ZJ8_DECK_CROSS", "aci": 4, "section": "HN-180x180", "role": "横向梁"},
    "deck_secondary_beam": {"cad_layer": "ZJ8_DECK_SECONDARY", "aci": 3, "section": "HN-120x120", "role": "平台次梁"},
    "cantilever_edge_beam": {"cad_layer": "ZJ8_CANTILEVER_EDGE", "aci": 3, "section": "HN-120x120", "role": "外挑边梁"},
    "platform_brace": {"cad_layer": "ZJ8_PLATFORM_BRACE", "aci": 1, "section": "L90x90", "role": "平台平面支撑"},
    "column": {"cad_layer": "ZJ8_COLUMN", "aci": 8, "section": "HM-240x240", "role": "柱子"},
    "transverse_column_brace": {"cad_layer": "ZJ8_TRANSVERSE_BRACE", "aci": 1, "section": "L90x90", "role": "横向柱间支撑"},
    "longitudinal_support_brace": {"cad_layer": "ZJ8_LONGITUDINAL_BRACE", "aci": 1, "section": "L90x90", "role": "固定支架纵向支撑"},
    "truss_upper_chord": {"cad_layer": "ZJ8_TRUSS_UPPER", "aci": 5, "section": "HN-180x180", "role": "桁架上弦"},
    "truss_lower_chord": {"cad_layer": "ZJ8_TRUSS_LOWER", "aci": 5, "section": "HN-180x180", "role": "桁架下弦"},
    "truss_vertical": {"cad_layer": "ZJ8_TRUSS_VERTICAL", "aci": 5, "section": "L90x90", "role": "桁架竖腹杆"},
    "truss_diagonal": {"cad_layer": "ZJ8_TRUSS_DIAGONAL", "aci": 5, "section": "L90x90", "role": "桁架斜腹杆"},
}

SECTION_SIZES: dict[str, tuple[float, float]] = {
    "HM-220x220": (220.0, 220.0),
    "HN-180x180": (180.0, 180.0),
    "HN-120x120": (120.0, 120.0),
    "HM-240x240": (240.0, 240.0),
    "L90x90": (90.0, 90.0),
}


@dataclass
class TrestleParameters:
    support_spans: list[float] = field(default_factory=lambda: [0.0, 6000.0, 6000.0, 13000.0, 12000.0])
    support_width: float = 3200.0
    start_elevation: float = 3200.0
    slope_deg: float = 4.0
    equipment_width: float = 2150.0
    y_origin: float = 0.0
    left_cantilever: float = 1125.0
    right_cantilever: float = 1625.0
    start_cantilever_length: float = 1400.0
    end_cantilever_length: float = 0.0
    platform_segment_limit: float = 3000.0
    column_segment_limit: float = 5000.0
    fixed_support_span_range: tuple[float, float] = (2000.0, 6500.0)
    truss_span_range: tuple[float, float] = (12000.0, 40000.0)
    truss_segment_limit: float = 2000.0
    truss_depth: float = 1800.0
    platform_scheme: str = "double-side-cantilever"
    brace_scheme: str = "alternating-x"
    truss_scheme: str = "warren-with-verticals"
    segment_spans: list[float] = field(default_factory=list)

    @property
    def support_stations(self) -> list[float]:
        if not self.support_spans:
            return [0.0]
        if abs(self.support_spans[0]) < 1e-9:
            stations = [0.0]
            for span in self.support_spans[1:]:
                stations.append(stations[-1] + span)
            return stations
        stations = [0.0]
        for span in self.support_spans:
            stations.append(stations[-1] + span)
        return stations

    @classmethod
    def from_csv(cls, path: str | Path) -> "TrestleParameters":
        data: dict[str, str] = {}
        support_spans: list[float] = []
        with Path(path).open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                key = (row.get("key") or row.get("参数") or "").strip()
                value = (row.get("value") or row.get("数值") or "").strip()
                if key:
                    data[key] = value
                span = row.get("support_span") or row.get("支架间距")
                if span not in (None, ""):
                    support_spans.append(float(span))
        return cls.from_mapping(data, support_spans=support_spans or None)

    @classmethod
    def from_xlsx(cls, path: str | Path) -> "TrestleParameters":
        from openpyxl import load_workbook

        workbook = load_workbook(path, data_only=True)
        sheet = workbook.active
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
        data: dict[str, str] = {}
        support_spans: list[float] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_map = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
            key = row_map.get("key") or row_map.get("参数")
            value = row_map.get("value") or row_map.get("数值")
            if key is not None and value is not None:
                data[str(key).strip()] = str(value).strip()
            span = row_map.get("support_span") or row_map.get("支架间距")
            if span not in (None, ""):
                support_spans.append(float(span))
        return cls.from_mapping(data, support_spans=support_spans or None)

    @classmethod
    def from_mapping(cls, data: dict[str, object], support_spans: list[float] | None = None) -> "TrestleParameters":
        aliases = {
            "支架宽度": "support_width",
            "平台宽度": "support_width",
            "起点高度": "start_elevation",
            "起点标高": "start_elevation",
            "倾角": "slope_deg",
            "平台倾角": "slope_deg",
            "设备宽度": "equipment_width",
            "左侧悬挑距离": "left_cantilever",
            "右侧悬挑距离": "right_cantilever",
            "始端悬挑长度": "start_cantilever_length",
            "末端悬挑长度": "end_cantilever_length",
            "单跨分格控制长度": "platform_segment_limit",
            "平台最大分段": "platform_segment_limit",
            "平台最大分段长度": "platform_segment_limit",
            "柱间支撑控制高度": "column_segment_limit",
            "柱最大分段": "column_segment_limit",
            "柱最大分段长度": "column_segment_limit",
            "桁架分段长度": "truss_segment_limit",
            "桁架高度": "truss_depth",
            "普通支架柱跨范围": "fixed_support_span_range",
            "固定支架柱跨范围": "fixed_support_span_range",
            "桁架柱跨范围": "truss_span_range",
            "支架里程": "support_stations",
            "支架定位": "support_stations",
            "支架间距": "support_spans",
            "拼接分段": "segment_spans",
            "分缝分段": "segment_spans",
            "串联拼装分段": "segment_spans",
        }
        kwargs: dict[str, object] = {}
        for key, value in data.items():
            attr = aliases.get(str(key), str(key))
            if attr == "support_stations":
                kwargs["support_spans"] = spans_from_stations(parse_float_list(value))
            elif attr in cls.__dataclass_fields__:
                if attr.endswith("_scheme"):
                    kwargs[attr] = str(value)
                elif attr in {"support_spans", "segment_spans"}:
                    kwargs[attr] = parse_float_list(value)
                elif attr.endswith("_range"):
                    kwargs[attr] = parse_float_range(value)
                else:
                    kwargs[attr] = float(value)
        if support_spans is not None:
            kwargs["support_spans"] = support_spans
        return cls(**kwargs)

    def to_dict(self) -> dict[str, object]:
        data = asdict(self)
        data["support_stations"] = self.support_stations
        return data


@dataclass(frozen=True)
class BaselineMember:
    id: str
    category: str
    start: Point3
    end: Point3
    section: str
    cad_layer: str
    aci: int
    segment: str
    scheme: str

    @property
    def length(self) -> float:
        return math.dist(self.start, self.end)

    def as_row(self) -> dict[str, object]:
        return {
            "id": self.id,
            "category": self.category,
            "role": LAYER_TABLE[self.category]["role"],
            "cad_layer": self.cad_layer,
            "aci": self.aci,
            "section": self.section,
            "segment": self.segment,
            "scheme": self.scheme,
            "start_x": round(self.start[0], 6),
            "start_y": round(self.start[1], 6),
            "start_z": round(self.start[2], 6),
            "end_x": round(self.end[0], 6),
            "end_y": round(self.end[1], 6),
            "end_z": round(self.end[2], 6),
            "length": round(self.length, 6),
        }


def build_members(params: TrestleParameters | None = None) -> list[BaselineMember]:
    p = params or TrestleParameters()
    stations = p.support_stations
    segment_ranges = compute_segment_ranges(p)
    members: list[BaselineMember] = []
    counter: Counter[str] = Counter()

    def pt(station: float, y: float, z_offset: float = 0.0) -> Point3:
        return (station, p.y_origin + y, p.start_elevation + math.tan(math.radians(p.slope_deg)) * station + z_offset)

    def segment_for_station(station: float) -> str:
        for index, (start, end) in enumerate(segment_ranges, start=1):
            if start - 1e-6 <= station <= end + 1e-6:
                return f"P{index}_{int(start)}_{int(end)}"
        return "P1"

    def add(category: str, start: Point3, end: Point3, segment: str, scheme: str = "") -> None:
        if math.dist(start, end) <= 1e-6:
            return
        counter[category] += 1
        info = LAYER_TABLE[category]
        members.append(
            BaselineMember(
                id=f"{category}_{counter[category]:04d}",
                category=category,
                start=start,
                end=end,
                section=str(info["section"]),
                cad_layer=str(info["cad_layer"]),
                aci=int(info["aci"]),
                segment=segment,
                scheme=scheme,
            )
        )

    half = p.support_width / 2.0
    left_edge = -p.equipment_width / 2.0 - p.left_cantilever
    right_edge = p.equipment_width / 2.0 + p.right_cantilever
    left_main = -half
    right_main = half
    left_inner = -p.equipment_width / 2.0
    right_inner = p.equipment_width / 2.0

    # Support columns and transverse bracing. Column subdivision follows the tutorial's limit-equalize card.
    for index, station in enumerate(stations):
        top_z = pt(station, 0.0)[2]
        for y in (left_main, right_main):
            levels = equal_subdivision_values(0.0, top_z, p.column_segment_limit)
            for z0, z1 in zip(levels[:-1], levels[1:]):
                add("column", (station, p.y_origin + y, z0), (station, p.y_origin + y, z1), f"Z{index + 1}", p.brace_scheme)
        support_segment = f"{segment_for_station(station)}:Z{index + 1}"
        add("deck_cross_beam", pt(station, left_edge), pt(station, right_edge), support_segment, p.platform_scheme)
        add("transverse_column_brace", (station, p.y_origin + left_main, 0.0), pt(station, right_main), support_segment, p.brace_scheme)
        add("transverse_column_brace", (station, p.y_origin + right_main, 0.0), pt(station, left_main), support_segment, p.brace_scheme)

    # Start/end cantilever platform edges.
    route_start = stations[0] - p.start_cantilever_length
    route_end = stations[-1] + p.end_cantilever_length
    if p.start_cantilever_length > 0:
        for y in (left_edge, right_edge):
            add("cantilever_edge_beam", pt(route_start, y), pt(stations[0], y), "START_CANTILEVER", p.platform_scheme)
    if p.end_cantilever_length > 0:
        for y in (left_edge, right_edge):
            add("cantilever_edge_beam", pt(stations[-1], y), pt(route_end, y), "END_CANTILEVER", p.platform_scheme)

    for bay_index, (a, b) in enumerate(zip(stations[:-1], stations[1:]), start=1):
        span = b - a
        segment = f"{segment_for_station((a + b) / 2.0)}:S{bay_index}_{int(a)}_{int(b)}"
        bay_stations = equal_subdivision_values(a, b, p.platform_segment_limit)

        # Main beams along support lines.
        for y in (left_main, right_main):
            for x0, x1 in zip(bay_stations[:-1], bay_stations[1:]):
                add("deck_main_beam", pt(x0, y), pt(x1, y), segment, p.platform_scheme)

        # Cross beams at all division points. Interior endpoints are preserved for bracing.
        for x in bay_stations[1:-1]:
            add("deck_cross_beam", pt(x, left_edge), pt(x, right_edge), segment, p.platform_scheme)

        # Longitudinal secondary and cantilever edge beams.
        for y in (left_inner, right_inner):
            for x0, x1 in zip(bay_stations[:-1], bay_stations[1:]):
                add("deck_secondary_beam", pt(x0, y), pt(x1, y), segment, p.platform_scheme)
        for y in (left_edge, right_edge):
            for x0, x1 in zip(bay_stations[:-1], bay_stations[1:]):
                add("cantilever_edge_beam", pt(x0, y), pt(x1, y), segment, p.platform_scheme)

        # Platform in-plane braces use the full endpoint list; this mirrors the tutorial's "do not drop endpoints yet" note.
        for i, (x0, x1) in enumerate(zip(bay_stations[:-1], bay_stations[1:])):
            if p.brace_scheme == "single-diagonal":
                add("platform_brace", pt(x0, left_main), pt(x1, right_main), segment, p.brace_scheme)
            elif i % 2 == 0:
                add("platform_brace", pt(x0, left_main), pt(x1, right_main), segment, p.brace_scheme)
            else:
                add("platform_brace", pt(x0, right_main), pt(x1, left_main), segment, p.brace_scheme)

        if in_range(span, p.fixed_support_span_range):
            add("longitudinal_support_brace", (a, p.y_origin + left_main, 0.0), pt(b, left_main), segment, p.brace_scheme)
            add("longitudinal_support_brace", (a, p.y_origin + right_main, 0.0), pt(b, right_main), segment, p.brace_scheme)

        if in_range(span, p.truss_span_range):
            truss_stations = equal_subdivision_values(a, b, p.truss_segment_limit)
            for y in (left_main, right_main):
                upper = [pt(x, y) for x in truss_stations]
                lower = [pt(x, y, -p.truss_depth) for x in truss_stations]
                for u0, u1 in zip(upper[:-1], upper[1:]):
                    add("truss_upper_chord", u0, u1, segment, p.truss_scheme)
                for l0, l1 in zip(lower[:-1], lower[1:]):
                    add("truss_lower_chord", l0, l1, segment, p.truss_scheme)
                for u, l in zip(upper, lower):
                    add("truss_vertical", u, l, segment, p.truss_scheme)
                for i in range(len(truss_stations) - 1):
                    if p.truss_scheme == "pratt":
                        add("truss_diagonal", lower[i], upper[i + 1], segment, p.truss_scheme)
                    elif p.truss_scheme == "x-braced":
                        add("truss_diagonal", lower[i], upper[i + 1], segment, p.truss_scheme)
                        add("truss_diagonal", upper[i], lower[i + 1], segment, p.truss_scheme)
                    else:
                        if i % 2 == 0:
                            add("truss_diagonal", lower[i], upper[i + 1], segment, p.truss_scheme)
                        else:
                            add("truss_diagonal", upper[i], lower[i + 1], segment, p.truss_scheme)

    return members


def tutorial_sample_parameters(slope_deg: float = 0.0) -> TrestleParameters:
    return TrestleParameters(
        support_spans=[0.0, 12000.0, 12000.0, 12000.0, 3000.0, 20520.0, 3000.0, 12000.0, 3000.0, 39000.0, 3000.0],
        support_width=5750.0,
        start_elevation=17000.0,
        slope_deg=slope_deg,
        equipment_width=2150.0,
        left_cantilever=1125.0,
        right_cantilever=1625.0,
        start_cantilever_length=1400.0,
        end_cantilever_length=0.0,
        platform_segment_limit=3000.0,
        column_segment_limit=5000.0,
        fixed_support_span_range=(2000.0, 6000.0),
        truss_span_range=(13000.0, 40000.0),
        truss_segment_limit=2000.0,
        truss_depth=1800.0,
        platform_scheme="tutorial-double-side-cantilever",
        brace_scheme="alternating-x",
        truss_scheme="warren-with-verticals",
        segment_spans=[39000.0, 38520.0, 42000.0],
    )


def compute_segment_ranges(params: TrestleParameters) -> list[tuple[float, float]]:
    total = params.support_stations[-1] if params.support_stations else 0.0
    if not params.segment_spans:
        return [(0.0, total)]
    ranges: list[tuple[float, float]] = []
    cursor = 0.0
    for span in params.segment_spans:
        start = cursor
        end = min(total, cursor + span)
        ranges.append((start, end))
        cursor = end
        if cursor >= total - 1e-6:
            break
    if not ranges or ranges[-1][1] < total - 1e-6:
        ranges.append((cursor, total))
    return ranges


def segment_label(index: int, start: float, end: float) -> str:
    return f"P{index}_{int(round(start))}_{int(round(end))}"


def split_members_by_segment(members: list[BaselineMember], params: TrestleParameters) -> dict[str, list[BaselineMember]]:
    ranges = compute_segment_ranges(params)
    grouped: dict[str, list[BaselineMember]] = {segment_label(i, start, end): [] for i, (start, end) in enumerate(ranges, start=1)}
    for member in members:
        midpoint = (member.start[0] + member.end[0]) / 2.0
        if midpoint < ranges[0][0]:
            grouped[segment_label(1, ranges[0][0], ranges[0][1])].append(member)
            continue
        if midpoint > ranges[-1][1]:
            grouped[segment_label(len(ranges), ranges[-1][0], ranges[-1][1])].append(member)
            continue
        for i, (start, end) in enumerate(ranges, start=1):
            if start - 1e-6 <= midpoint <= end + 1e-6:
                grouped[segment_label(i, start, end)].append(member)
                break
    return grouped


def _write_dxf(members: list[BaselineMember], path: str | Path, params: TrestleParameters | None = None, include_splices: bool = False) -> Path:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    lines: list[str] = []

    def pair(code: int, value: object) -> None:
        lines.append(str(code))
        lines.append(str(value))

    pair(0, "SECTION")
    pair(2, "HEADER")
    pair(9, "$INSUNITS")
    pair(70, 4)
    pair(0, "ENDSEC")
    pair(0, "SECTION")
    pair(2, "TABLES")
    pair(0, "TABLE")
    pair(2, "LAYER")
    layer_count = len(LAYER_TABLE) + (2 if include_splices and params is not None else 0)
    pair(70, layer_count)
    for info in LAYER_TABLE.values():
        pair(0, "LAYER")
        pair(2, info["cad_layer"])
        pair(70, 0)
        pair(62, info["aci"])
        pair(6, "CONTINUOUS")
    if include_splices and params is not None:
        for layer, color in (("ZJ8_SPLICE_JOINT", 2), ("ZJ8_SPLICE_TEXT", 2)):
            pair(0, "LAYER")
            pair(2, layer)
            pair(70, 0)
            pair(62, color)
            pair(6, "CONTINUOUS")
    pair(0, "ENDTAB")
    pair(0, "ENDSEC")
    pair(0, "SECTION")
    pair(2, "ENTITIES")
    for member in members:
        pair(0, "LINE")
        pair(8, member.cad_layer)
        pair(10, round(member.start[0], 6))
        pair(20, round(member.start[1], 6))
        pair(30, round(member.start[2], 6))
        pair(11, round(member.end[0], 6))
        pair(21, round(member.end[1], 6))
        pair(31, round(member.end[2], 6))
    if include_splices and params is not None:
        half = params.support_width / 2.0
        z0 = 0.0
        z1 = params.start_elevation + math.tan(math.radians(params.slope_deg)) * params.support_stations[-1] + 4000.0
        for index, (_start, end) in enumerate(compute_segment_ranges(params)[:-1], start=1):
            x = end
            pair(0, "LINE")
            pair(8, "ZJ8_SPLICE_JOINT")
            pair(10, x)
            pair(20, -half - 800.0)
            pair(30, z0)
            pair(11, x)
            pair(21, half + 800.0)
            pair(31, z1)
            pair(0, "TEXT")
            pair(8, "ZJ8_SPLICE_TEXT")
            pair(10, x + 250.0)
            pair(20, half + 1100.0)
            pair(30, z1)
            pair(40, 450.0)
            pair(1, f"SPLICE P{index}/P{index + 1}")
    pair(0, "ENDSEC")
    pair(0, "EOF")
    target.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return target


def export_segment_joints_dxf(params: TrestleParameters, path: str | Path) -> Path:
    return _write_dxf([], path, params=params, include_splices=True)


def equal_subdivision_values(start: float, end: float, limit: float) -> list[float]:
    length = abs(end - start)
    if length <= 1e-9:
        return [start]
    count = max(1, math.ceil(length / limit))
    return [start + (end - start) * i / count for i in range(count + 1)]


def trim_head_tail(values: list[float]) -> list[float]:
    return list(values[1:-1]) if len(values) > 2 else []


def in_range(value: float, bounds: tuple[float, float]) -> bool:
    return bounds[0] <= abs(value) <= bounds[1]


def optimize_concept(params: TrestleParameters) -> dict[str, object]:
    candidates = []
    for platform_limit in (2500.0, 3000.0, 3500.0):
        for column_limit in (4000.0, 5000.0):
            for truss_depth in (1500.0, 1800.0, 2200.0):
                base = asdict(params)
                base.update(
                    {
                        "platform_segment_limit": platform_limit,
                        "column_segment_limit": column_limit,
                        "truss_depth": truss_depth,
                    }
                )
                candidate = TrestleParameters(**base)
                members = build_members(candidate)
                total_length = sum(m.length for m in members)
                member_count = len(members)
                score = total_length * 0.001 + member_count * 0.25 + truss_depth * 0.0005
                candidates.append(
                    {
                        "platform_segment_limit": platform_limit,
                        "column_segment_limit": column_limit,
                        "truss_depth": truss_depth,
                        "member_count": member_count,
                        "total_member_length": round(total_length, 3),
                        "concept_score": round(score, 3),
                    }
                )
    candidates.sort(key=lambda item: item["concept_score"])
    return {"note": "Concept-only ranking; not a structural calculation.", "best": candidates[0], "candidates": candidates[:8]}


def summarize_members(members: Iterable[BaselineMember]) -> list[dict[str, object]]:
    rows = [member.as_row() for member in members]
    summary: dict[tuple[str, str, str], dict[str, object]] = {}
    for row in rows:
        key = (str(row["category"]), str(row["cad_layer"]), str(row["section"]))
        item = summary.setdefault(
            key,
            {
                "category": row["category"],
                "role": row["role"],
                "cad_layer": row["cad_layer"],
                "section": row["section"],
                "count": 0,
                "total_length": 0.0,
            },
        )
        item["count"] = int(item["count"]) + 1
        item["total_length"] = float(item["total_length"]) + float(row["length"])
    return [
        {**item, "total_length": round(float(item["total_length"]), 3)}
        for item in sorted(summary.values(), key=lambda value: str(value["category"]))
    ]


def export_baseline_csv(members: list[BaselineMember], path: str | Path) -> Path:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    rows = [member.as_row() for member in members]
    with target.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()) if rows else [])
        writer.writeheader()
        writer.writerows(rows)
    return target


def export_summary_csv(members: list[BaselineMember], path: str | Path) -> Path:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    rows = summarize_members(members)
    with target.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()) if rows else [])
        writer.writeheader()
        writer.writerows(rows)
    return target


def export_segment_summary_csv(members: list[BaselineMember], params: TrestleParameters, path: str | Path) -> Path:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    grouped = split_members_by_segment(members, params)
    rows: list[dict[str, object]] = []
    for label, segment_members in grouped.items():
        length = sum(member.length for member in segment_members)
        rows.append(
            {
                "segment": label,
                "member_count": len(segment_members),
                "total_member_length": round(length, 3),
                "start_station": label.split("_")[1],
                "end_station": label.split("_")[2],
            }
        )
    with target.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()) if rows else [])
        writer.writeheader()
        writer.writerows(rows)
    return target


def export_parameters_json(params: TrestleParameters, path: str | Path) -> Path:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(json.dumps(params.to_dict(), ensure_ascii=False, indent=2), encoding="utf-8")
    return target


def export_3d_baseline_dxf(
    members: list[BaselineMember],
    path: str | Path,
    params: TrestleParameters | None = None,
    include_splices: bool = False,
) -> Path:
    return _write_dxf(members, path, params=params, include_splices=include_splices)


def create_sample_parameter_csv(path: str | Path, params: TrestleParameters | None = None) -> Path:
    p = params or TrestleParameters()
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    rows = [
        {"key": "支架宽度", "value": p.support_width, "support_span": ""},
        {"key": "起点标高", "value": p.start_elevation, "support_span": ""},
        {"key": "平台倾角", "value": p.slope_deg, "support_span": ""},
        {"key": "设备宽度", "value": p.equipment_width, "support_span": ""},
        {"key": "左侧悬挑距离", "value": p.left_cantilever, "support_span": ""},
        {"key": "右侧悬挑距离", "value": p.right_cantilever, "support_span": ""},
        {"key": "始端悬挑长度", "value": p.start_cantilever_length, "support_span": ""},
        {"key": "末端悬挑长度", "value": p.end_cantilever_length, "support_span": ""},
        {"key": "单跨分格控制长度", "value": p.platform_segment_limit, "support_span": ""},
        {"key": "柱间支撑控制高度", "value": p.column_segment_limit, "support_span": ""},
        {"key": "桁架分段长度", "value": p.truss_segment_limit, "support_span": ""},
        {"key": "桁架高度", "value": p.truss_depth, "support_span": ""},
        {"key": "拼接分段", "value": ";".join(str(v) for v in p.segment_spans), "support_span": ""},
    ]
    for span in p.support_spans:
        rows.append({"key": "", "value": "", "support_span": span})
    with target.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["key", "value", "support_span"])
        writer.writeheader()
        writer.writerows(rows)
    return target


def create_sample_parameter_xlsx(path: str | Path, params: TrestleParameters | None = None) -> Path:
    from openpyxl import Workbook

    p = params or TrestleParameters()
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "参数统计表"
    sheet.append(["key", "value", "support_span", "说明"])
    for row in [
        ("支架宽度", p.support_width, "", "平台支架横向宽度"),
        ("起点标高", p.start_elevation, "", "起点高度"),
        ("平台倾角", p.slope_deg, "", "角度制"),
        ("设备宽度", p.equipment_width, "", "输送设备宽度"),
        ("左侧悬挑距离", p.left_cantilever, "", "由设备边到左平台边"),
        ("右侧悬挑距离", p.right_cantilever, "", "由设备边到右平台边"),
        ("始端悬挑长度", p.start_cantilever_length, "", "路线起点前外挑"),
        ("末端悬挑长度", p.end_cantilever_length, "", "路线末端后外挑"),
        ("单跨分格控制长度", p.platform_segment_limit, "", "平台等分上限"),
        ("柱间支撑控制高度", p.column_segment_limit, "", "柱段等分上限"),
        ("桁架分段长度", p.truss_segment_limit, "", "桁架腹杆分段"),
        ("桁架高度", p.truss_depth, "", "空间桁架高度"),
        ("拼接分段", ";".join(str(v) for v in p.segment_spans), "", "超长栈桥分缝段长"),
    ]:
        sheet.append(list(row))
    for span in p.support_spans:
        sheet.append(["", "", span, "支架间距"])
    workbook.save(target)
    return target
